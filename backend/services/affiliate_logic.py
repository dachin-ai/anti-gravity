import io
import pandas as pd
import numpy as np
import base64
import traceback
import math

NEGATIVE_METRICS = {'Refunded GMV', 'Refunded Items', 'Cancellation Rate'}

def clean_numeric_col(series):
    return pd.to_numeric(
        series.astype(str).str.replace('--', '0', regex=False).str.replace(',', '', regex=False),
        errors='coerce'
    ).fillna(0)

def clean_rp_col(series):
    cleaned = (
        series.astype(str)
        .str.replace('--', '0', regex=False)
        .str.replace('Rp', '', regex=False)
        .str.replace('.', '', regex=False)
        .str.replace(',', '', regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors='coerce').fillna(0)

def process_creator_file(df, prefix):
    col_mapping = {
        'Creator username': 'Creator Name',
        'Affiliate GMV': 'GMV',
        'Items sold': 'Item Sold',
        'Est. commission': 'Est. Commission',
        'Affiliate refunded GMV': 'Refunded GMV',
        'Affiliate items refunded': 'Refunded Items'
    }
    available_cols = [c for c in col_mapping.keys() if c in df.columns]
    df_clean = df[available_cols].rename(columns=col_mapping)
    for col in df_clean.columns:
        if col != 'Creator Name':
            df_clean[col] = clean_numeric_col(df_clean[col])
    
    if 'GMV' in df_clean.columns and 'Refunded GMV' in df_clean.columns:
        df_clean['NMV'] = df_clean['GMV'] - df_clean['Refunded GMV']
        df_clean['Cancellation Rate'] = np.where(
            df_clean['GMV'] > 0, df_clean['Refunded GMV'] / df_clean['GMV'], 0
        )
        
    rename_dict = {col: f"{col} ({prefix})" for col in df_clean.columns if col != 'Creator Name'}
    df_clean.rename(columns=rename_dict, inplace=True)
    return df_clean

def process_product_file(df, prefix):
    col_mapping = {
        'Product ID': 'Product ID',
        'Product name': 'Product Name',
        'GMV': 'GMV',
        'Items sold': 'Item Sold',
        'Est. commission': 'Est. Commission',
        'Refunded GMV': 'Refunded GMV',
        'Refunded items sold': 'Refunded Items'
    }
    available_cols = [c for c in col_mapping.keys() if c in df.columns]
    df_clean = df[available_cols].rename(columns=col_mapping)

    rp_cols = ['GMV', 'Est. Commission', 'Refunded GMV']
    for col in rp_cols:
        if col in df_clean.columns:
            df_clean[col] = clean_rp_col(df_clean[col])

    num_cols = ['Item Sold', 'Refunded Items']
    for col in num_cols:
        if col in df_clean.columns:
            df_clean[col] = clean_numeric_col(df_clean[col])

    if 'Product ID' in df_clean.columns:
        df_clean['Product ID'] = df_clean['Product ID'].astype(str).str.strip()

    if 'GMV' in df_clean.columns and 'Refunded GMV' in df_clean.columns:
        df_clean['NMV'] = df_clean['GMV'] - df_clean['Refunded GMV']
        df_clean['Cancellation Rate'] = np.where(
            df_clean['GMV'] > 0, df_clean['Refunded GMV'] / df_clean['GMV'], 0
        )

    rename_dict = {col: f"{col} ({prefix})" for col in df_clean.columns if col not in ('Product ID', 'Product Name')}
    df_clean.rename(columns=rename_dict, inplace=True)
    return df_clean

def build_styled_excel(df_export, a_missing, b_missing, name_col, sheet_name):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0C2461', end_color='0C2461', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=10)
    data_font_bold = Font(name='Calibri', size=10, bold=True)
    num_align = Alignment(horizontal='right', vertical='center')
    text_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    green_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    green_font = Font(name='Calibri', size=10, bold=True, color='166534')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    red_font = Font(name='Calibri', size=10, bold=True, color='991B1B')
    na_font = Font(name='Calibri', size=10, color='9CA3AF')
    
    thin_border = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
                         top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'))
    header_border = Border(left=Side(style='thin', color='1A3A7A'), right=Side(style='thin', color='1A3A7A'),
                           top=Side(style='thin', color='1A3A7A'), bottom=Side(style='thin', color='1A3A7A'))

    cols = list(df_export.columns)
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.freeze_panes = 'A2'

    growth_cols_idx = set()
    cancellation_cols_idx = set()
    pct_cols_idx = set()
    numeric_data_cols_idx = set()
    growth_metric_map = {}

    for c_idx, col_name in enumerate(cols, 1):
        if 'Growth' in col_name:
            growth_cols_idx.add(c_idx)
            pct_cols_idx.add(c_idx)
            metric_name = col_name.replace(' Growth', '')
            growth_metric_map[c_idx] = metric_name
        elif 'Cancellation Rate' in col_name:
            cancellation_cols_idx.add(c_idx)
            pct_cols_idx.add(c_idx)
        elif col_name not in ['Rank', name_col, 'Product ID', 'Product Name', '_key']:
            numeric_data_cols_idx.add(c_idx)

    for r_idx, (_, row) in enumerate(df_export.iterrows(), 2):
        is_a_miss = a_missing[r_idx - 2]
        is_b_miss = b_missing[r_idx - 2]

        for c_idx, col_name in enumerate(cols, 1):
            val = row[col_name]
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border

            if col_name == 'Rank':
                cell.value = int(val) if not pd.isna(val) else val
                cell.font = data_font_bold
                cell.alignment = center_align
                continue

            if col_name in (name_col, 'Product ID', 'Product Name', '_key'):
                cell.value = str(val) if val is not None else ""
                cell.font = data_font
                cell.alignment = text_align
                continue

            is_col_a = '(A)' in col_name
            is_col_b = '(B)' in col_name
            is_growth = 'Growth' in col_name
            show_na = (is_col_a and is_a_miss) or (is_col_b and is_b_miss) or (is_growth and (is_a_miss or is_b_miss))

            if show_na:
                cell.value = '-'
                cell.font = na_font
                cell.alignment = center_align
                continue

            if c_idx in growth_cols_idx:
                try:
                    v = float(val)
                    if math.isnan(v) or math.isinf(v):
                        cell.value = '-'
                    else:
                        cell.value = v
                        cell.number_format = '+0.00%;-0.00%;0.00%'
                        cell.alignment = num_align
                        metric_name = growth_metric_map.get(c_idx, '')
                        is_neg_metric = metric_name in NEGATIVE_METRICS
                        if v > 0:
                            if is_neg_metric:
                                cell.fill = red_fill; cell.font = red_font
                            else:
                                cell.fill = green_fill; cell.font = green_font
                        elif v < 0:
                            if is_neg_metric:
                                cell.fill = green_fill; cell.font = green_font
                            else:
                                cell.fill = red_fill; cell.font = red_font
                        else:
                            cell.font = data_font
                except:
                    cell.value = '-'
                continue

            if c_idx in cancellation_cols_idx:
                try:
                    cell.value = float(val)
                    cell.number_format = '0.00%'
                except:
                    cell.value = 0
                cell.alignment = num_align
                cell.font = data_font
                continue

            if c_idx in numeric_data_cols_idx:
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0'
                except:
                    cell.value = 0
                cell.alignment = num_align
                cell.font = data_font
                continue

            cell.value = val
            cell.font = data_font

    even_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    for r_idx in range(2, ws.max_row + 1):
        if r_idx % 2 == 0:
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if cell.fill.start_color.index in ('00000000', 0) or cell.fill.patternType is None:
                    cell.fill = even_fill

    for c_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c_idx)
        ws.column_dimensions[col_letter].width = 20

    ws.row_dimensions[1].height = 30
    return wb

def analyze_affiliates(file_content_a: bytes, is_csv_a: bool, file_content_b: bytes, is_csv_b: bool, mode: str):
    df_a = pd.read_csv(io.BytesIO(file_content_a)) if is_csv_a else pd.read_excel(io.BytesIO(file_content_a))
    df_b = pd.read_csv(io.BytesIO(file_content_b)) if is_csv_b else pd.read_excel(io.BytesIO(file_content_b))

    if mode == 'Creator':
        name_col = 'Creator Name'
        merge_key = 'Creator Name'
        df_old_proc = process_creator_file(df_a, 'A')
        df_new_proc = process_creator_file(df_b, 'B')
        df_merge = pd.merge(df_old_proc, df_new_proc, on=merge_key, how='outer')
        df_merge['_a_missing'] = df_merge['GMV (A)'].isna()
        df_merge['_b_missing'] = df_merge['GMV (B)'].isna()
        numeric_cols = [c for c in df_merge.columns if c != merge_key and not c.startswith('_')]
        df_merge[numeric_cols] = df_merge[numeric_cols].fillna(0)
        df_merge = df_merge[~((df_merge['GMV (A)'] == 0) & (df_merge['GMV (B)'] == 0))].copy()
    else:
        name_col = 'Product ID'
        df_old_proc = process_product_file(df_a, 'A')
        df_new_proc = process_product_file(df_b, 'B')
        df_merge = pd.merge(df_old_proc, df_new_proc, on='Product ID', how='outer', suffixes=('_old', '_new'))
        
        if 'Product Name_new' in df_merge.columns and 'Product Name_old' in df_merge.columns:
            df_merge['Product Name'] = df_merge['Product Name_new'].fillna(df_merge['Product Name_old'])
            df_merge.drop(columns=['Product Name_old', 'Product Name_new'], inplace=True)
        elif 'Product Name' not in df_merge.columns:
            if 'Product Name_new' in df_merge.columns:
                df_merge.rename(columns={'Product Name_new': 'Product Name'}, inplace=True)
            elif 'Product Name_old' in df_merge.columns:
                df_merge.rename(columns={'Product Name_old': 'Product Name'}, inplace=True)

        df_merge['_a_missing'] = df_merge['GMV (A)'].isna()
        df_merge['_b_missing'] = df_merge['GMV (B)'].isna()
        numeric_cols = [c for c in df_merge.columns if c not in ('Product ID', 'Product Name') and not c.startswith('_')]
        df_merge[numeric_cols] = df_merge[numeric_cols].fillna(0)

    metrics = ['GMV', 'Item Sold', 'Est. Commission', 'Refunded GMV', 'Refunded Items', 'NMV', 'Cancellation Rate']
    for m in metrics:
        oc, nc, gc = f"{m} (A)", f"{m} (B)", f"{m} Growth"
        if oc in df_merge.columns and nc in df_merge.columns:
            df_merge[gc] = np.where(
                df_merge[oc] > 0,
                (df_merge[nc] - df_merge[oc]) / df_merge[oc],
                np.where(df_merge[nc] > 0, 1.0, 0.0)
            )

    df_merge.sort_values(by='GMV (B)', ascending=False, inplace=True)
    df_merge.reset_index(drop=True, inplace=True)
    df_merge.insert(0, 'Rank', df_merge.index + 1)

    a_missing = df_merge['_a_missing'].values
    b_missing = df_merge['_b_missing'].values
    
    # Calculate top5s
    def get_top5(col, fmt, name_key):
        if col not in df_merge.columns: return []
        valid_df = df_merge[df_merge[col] > 0] if fmt != 'percent' else df_merge
        top = valid_df.nlargest(5, col)
        
        results = []
        for _, row in top.iterrows():
            if math.isnan(row[col]) or math.isinf(row[col]): continue
            results.append({
                "name": str(row[name_key]),
                "value": float(row[col]),
                "fmt": fmt
            })
        return results

    top5_data = {
        "gmv_b": get_top5('GMV (B)', 'currency', name_col),
        "nmv_b": get_top5('NMV (B)', 'currency', name_col),
        "item_b": get_top5('Item Sold (B)', 'number', name_col),
        "gmv_growth": get_top5('GMV Growth', 'percent', name_col),
        "nmv_growth": get_top5('NMV Growth', 'percent', name_col),
        "item_growth": get_top5('Item Sold Growth', 'percent', name_col),
    }

    total_gmv_a = float(df_merge['GMV (A)'].sum()) if 'GMV (A)' in df_merge.columns else 0
    total_gmv_b = float(df_merge['GMV (B)'].sum()) if 'GMV (B)' in df_merge.columns else 0
    gmv_change = ((total_gmv_b - total_gmv_a) / total_gmv_a) if total_gmv_a > 0 else 0

    total_nmv_a = float(df_merge['NMV (A)'].sum()) if 'NMV (A)' in df_merge.columns else 0
    total_nmv_b = float(df_merge['NMV (B)'].sum()) if 'NMV (B)' in df_merge.columns else 0
    nmv_change = ((total_nmv_b - total_nmv_a) / total_nmv_a) if total_nmv_a > 0 else 0

    total_items_a = float(df_merge['Item Sold (A)'].sum()) if 'Item Sold (A)' in df_merge.columns else 0
    total_items_b = float(df_merge['Item Sold (B)'].sum()) if 'Item Sold (B)' in df_merge.columns else 0

    summary = {
        "total_entities": len(df_merge),
        "new_entities": int(df_merge['_a_missing'].sum()),
        "lost_entities": int(df_merge['_b_missing'].sum()),
        "total_gmv_b": total_gmv_b, "gmv_change": gmv_change,
        "total_nmv_b": total_nmv_b, "nmv_change": nmv_change,
        "total_items_b": total_items_b, "total_items_a": total_items_a
    }

    # Formatting table JSON
    if mode == 'Product':
        ordered_cols = ['Rank', 'Product ID', 'Product Name']
    else:
        ordered_cols = ['Rank', name_col]
        
    for m in metrics:
        ordered_cols += [f"{m} (A)", f"{m} (B)", f"{m} Growth"]
        
    ordered_cols = [c for c in ordered_cols if c in df_merge.columns]
    
    # We must sanitize the dataframe for JSON (nan/inf won't serialize cleanly)
    df_table = df_merge[ordered_cols].copy()
    df_export = df_merge[ordered_cols].copy()
    
    # Ensure all float columns are cleansed for JSON mapping
    def sanitize_val(val):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return None
        return val

    table_rows = []
    for idx, row in df_table.iterrows():
        r_dict = {'key': idx, '_a_missing': bool(a_missing[idx]), '_b_missing': bool(b_missing[idx])}
        for col in ordered_cols:
            r_dict[col] = sanitize_val(row[col])
        table_rows.append(r_dict)

    # Export
    buffer = io.BytesIO()
    sheet_name = f'{"Product" if mode == "Product" else "Creator"} Comparison'
    wb = build_styled_excel(df_export, a_missing, b_missing, name_col, sheet_name)
    wb.save(buffer)
    b64_str = base64.b64encode(buffer.getvalue()).decode('utf-8')

    return {
        "summary": summary,
        "top5": top5_data,
        "table": table_rows,
        "metrics": metrics,
        "file_base64": b64_str
    }
