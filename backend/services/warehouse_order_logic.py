import pandas as pd
import numpy as np
import io
import math
from typing import Dict, List, Any
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


def process_warehouse_order(
    aov: float,
    total_days: int,
    platforms: List[Dict],         # [{name, number, target}]
    warehouses: List[Dict],        # [{name, number, proportion}]
    events: List[Dict],            # [{name, dates_str, proportion}]
) -> Dict[str, Any]:
    """
    Core calculation for Warehouse Order Estimator.
    Returns summary, matrix (as list-of-dicts), column structure, and Excel bytes.
    """
    # --- Sort & clean ---
    plat_df = pd.DataFrame(platforms).dropna(subset=['name'])
    plat_df['number'] = pd.to_numeric(plat_df['number'], errors='coerce').fillna(99)
    plat_df['target'] = pd.to_numeric(plat_df['target'], errors='coerce')
    plat_df = plat_df.sort_values('number').reset_index(drop=True)

    wh_df = pd.DataFrame(warehouses).dropna(subset=['name'])
    wh_df['number'] = pd.to_numeric(wh_df['number'], errors='coerce').fillna(99)
    wh_df['proportion'] = pd.to_numeric(wh_df['proportion'], errors='coerce').fillna(0)
    wh_df = wh_df.sort_values('number').reset_index(drop=True)

    total_wh_prop = wh_df['proportion'].sum()

    # --- Event proportion per day ---
    day_proportions: Dict[int, float] = {}
    assigned_days: set = set()
    summary_details = []

    for ev in events:
        name = ev.get('name', '')
        date_str = str(ev.get('dates_str', ''))
        total_prop = float(ev.get('proportion', 0) or 0)
        if not date_str or date_str == 'nan':
            continue
        dates = [int(d.strip()) for d in date_str.split('+') if d.strip().isdigit()]
        if not dates:
            continue
        each_day_prop = total_prop / len(dates)
        for d in dates:
            if 1 <= d <= total_days:
                day_proportions[d] = each_day_prop
                assigned_days.add(d)
        summary_details.append({
            'Event': name,
            'Days Count': len(dates),
            'Total Prop (%)': f"{total_prop}%",
            'Each Day (%)': f"{each_day_prop:.2f}%"
        })

    # Regular days
    used_prop = sum(day_proportions.values())
    reg_prop_total = 100.0 - used_prop
    all_days = set(range(1, total_days + 1))
    reg_days = list(all_days - assigned_days)
    reg_prop_daily = 0.0
    if reg_days:
        reg_prop_daily = reg_prop_total / len(reg_days)
        for d in reg_days:
            day_proportions[d] = reg_prop_daily
        summary_details.append({
            'Event': 'Regular Day',
            'Days Count': len(reg_days),
            'Total Prop (%)': f"{reg_prop_total:.2f}%",
            'Each Day (%)': f"{reg_prop_daily:.2f}%"
        })

    # --- Matrix calculation ---
    wh_names = wh_df['name'].tolist()
    plat_names = plat_df['name'].tolist()
    days_list = list(range(1, total_days + 1))

    # Structure: rows = warehouses, cols = (day, platform)
    # Returns as list of row dicts for JSON serialization
    matrix_rows = []
    matrix_raw = {}  # for Excel

    for _, w_row in wh_df.iterrows():
        w_name = w_row['name']
        w_prop = (w_row['proportion'] / total_wh_prop) if total_wh_prop > 0 else 0
        row_data = {'Warehouse': w_name}
        for d in days_list:
            day_prop = day_proportions.get(d, 0) / 100.0
            for _, p_row in plat_df.iterrows():
                p_name = p_row['name']
                p_target = p_row['target']
                col_key = f"D{d}|{p_name}"
                if pd.isna(p_target) or p_target <= 0 or w_row['proportion'] <= 0:
                    row_data[col_key] = None
                else:
                    val = math.ceil((float(p_target) * day_prop * w_prop) / float(aov))
                    row_data[col_key] = int(val)
        matrix_rows.append(row_data)

    # Column structure for frontend
    col_structure = []
    for d in days_list:
        for p_name in plat_names:
            col_structure.append({'day': d, 'platform': p_name, 'key': f"D{d}|{p_name}"})

    # --- Excel Export ---
    excel_bytes = _build_excel(matrix_rows, col_structure, wh_names, plat_names, days_list, aov, total_wh_prop, summary_details)

    return {
        'summary': summary_details,
        'aov': aov,
        'total_wh_prop': float(total_wh_prop),
        'platforms': plat_df[['name', 'target']].replace({np.nan: None}).to_dict(orient='records'),
        'warehouses': wh_df[['name', 'proportion']].to_dict(orient='records'),
        'matrix_rows': matrix_rows,
        'col_structure': col_structure,
        'days': days_list,
        'platform_names': plat_names,
        'warehouse_names': wh_names,
        'excel_b64': _b64(excel_bytes),
    }


def _b64(data: bytes) -> str:
    import base64
    return base64.b64encode(data).decode('utf-8')


def _build_excel(matrix_rows, col_structure, wh_names, plat_names, days_list, aov, total_wh_prop, summary_details) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Sheet 1: Estimation Matrix — manual MultiIndex build
        # Row 1: Day numbers (merged per platform count)
        # Row 2: Platform names
        # Row 3+: Warehouse data
        wb = writer.book
        ws_matrix = wb.create_sheet(title='Estimations')
        wb.active = ws_matrix

        navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=10)
        center = Alignment(horizontal="center", vertical="center")
        num_fmt = '#,##0'
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        n_plat = len(plat_names)

        # Header row 1: "Warehouse" + day labels (each spanning n_plat cols)
        ws_matrix.cell(1, 1, 'Warehouse').fill = navy_fill
        ws_matrix.cell(1, 1).font = white_font
        ws_matrix.cell(1, 1).alignment = center
        ws_matrix.cell(2, 1, 'Warehouse').fill = navy_fill
        ws_matrix.cell(2, 1).font = white_font
        ws_matrix.cell(2, 1).alignment = center

        col_offset = 2
        for d in days_list:
            start_col = col_offset
            end_col = col_offset + n_plat - 1
            ws_matrix.cell(1, start_col, f"Day {d}").fill = navy_fill
            ws_matrix.cell(1, start_col).font = white_font
            ws_matrix.cell(1, start_col).alignment = center
            if n_plat > 1:
                ws_matrix.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            for i, p_name in enumerate(plat_names):
                c = ws_matrix.cell(2, start_col + i, p_name)
                c.fill = navy_fill
                c.font = white_font
                c.alignment = center
            col_offset += n_plat

        # Data rows
        for r_idx, row_data in enumerate(matrix_rows):
            row_num = r_idx + 3
            wh_cell = ws_matrix.cell(row_num, 1, row_data['Warehouse'])
            wh_cell.fill = navy_fill
            wh_cell.font = white_font
            wh_cell.alignment = center
            col_offset = 2
            for cs in col_structure:
                val = row_data.get(cs['key'])
                cell = ws_matrix.cell(row_num, col_offset, val if val is not None else '')
                if val is not None:
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='right')
                col_offset += 1

        # Apply borders to all used cells
        for r in range(1, len(matrix_rows) + 3):
            for c in range(1, 2 + len(col_structure)):
                ws_matrix.cell(r, c).border = border

        # Auto-width - skip merged cells (they don't have column_letter)
        from openpyxl.cell.cell import MergedCell
        for col_cells in ws_matrix.columns:
            non_merged = [cell for cell in col_cells if not isinstance(cell, MergedCell)]
            if not non_merged:
                continue
            max_len = max((len(str(cell.value or '')) for cell in non_merged), default=8)
            ws_matrix.column_dimensions[non_merged[0].column_letter].width = min(max_len + 4, 20)

        # Sheet 2: Summary
        sum_df = pd.DataFrame(summary_details)
        sum_df.to_excel(writer, sheet_name='Summary', index=False)

    buf.seek(0)
    return buf.read()
