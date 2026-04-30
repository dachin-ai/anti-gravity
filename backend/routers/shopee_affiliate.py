from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from services.permission_guard import require_tool_access
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, String, case
from database import get_db
from typing import Optional
import io

from services.shopee_affiliate_logic import get_shopee_stores, process_and_save_upload
from models import ShopeeAffConversion, ShopeeAffProduct, ShopeeAffCreator, ProductPerformance, PidStoreMap

router = APIRouter()

# ─────────────────────────────────────────────────────────────────────────────
# STORES
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/stores", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def fetch_stores(db: Session = Depends(get_db)):
    stores = get_shopee_stores()
    # Fallback for local/dev when Google Sheet credentials are unavailable.
    # Keep UI usable by deriving store options from existing affiliate data.
    if not stores:
        store_ids = set()
        store_ids.update([r[0] for r in db.query(ShopeeAffConversion.store_id).filter(ShopeeAffConversion.store_id.isnot(None)).distinct().all() if r[0]])
        store_ids.update([r[0] for r in db.query(ShopeeAffProduct.store_id).filter(ShopeeAffProduct.store_id.isnot(None)).distinct().all() if r[0]])
        store_ids.update([r[0] for r in db.query(ShopeeAffCreator.store_id).filter(ShopeeAffCreator.store_id.isnot(None)).distinct().all() if r[0]])
        # Broader fallback: include known store universe from other Freemir datasets.
        store_ids.update([r[0] for r in db.query(ProductPerformance.store).filter(ProductPerformance.store.isnot(None)).distinct().all() if r[0]])
        store_ids.update([r[0] for r in db.query(PidStoreMap.store).filter(PidStoreMap.store.isnot(None)).distinct().all() if r[0]])
        store_ids = {sid.strip() for sid in store_ids if sid and sid.strip() and sid.strip() != "-"}
        stores = [{"code": sid, "name": sid} for sid in sorted(store_ids)]
    return {"stores": stores}

# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD (ETL)
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/upload", dependencies=[Depends(require_tool_access("affiliate_performance"))])
async def upload_shopee_csv(
    file: UploadFile = File(...),
    file_type: str = Form(...),
    store_id: str = Form(...),
    manual_date: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    valid_types = ['conversion', 'product', 'creator']
    if file_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid file type. Must be one of {valid_types}.")

    contents = await file.read()
    result = process_and_save_upload(db, contents, file.filename, file_type, store_id, manual_date)

    if not result.get("succeed"):
        raise HTTPException(status_code=400, detail=result.get("message"))
    return result

# ─────────────────────────────────────────────────────────────────────────────
# CHECKER MATRIX
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/checker-matrix", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def get_checker_matrix(month: str, year: str, db: Session = Depends(get_db)):
    from calendar import monthrange
    import datetime

    year_int  = int(year)
    month_int = int(month)
    _, num_days = monthrange(year_int, month_int)

    st_prod = db.query(ShopeeAffProduct.date, ShopeeAffProduct.store_id, func.count(ShopeeAffProduct.id)).filter(
        func.extract('year',  ShopeeAffProduct.date) == year_int,
        func.extract('month', ShopeeAffProduct.date) == month_int
    ).group_by(ShopeeAffProduct.date, ShopeeAffProduct.store_id).all()

    st_creator = db.query(ShopeeAffCreator.date, ShopeeAffCreator.store_id, func.count(ShopeeAffCreator.id)).filter(
        func.extract('year',  ShopeeAffCreator.date) == year_int,
        func.extract('month', ShopeeAffCreator.date) == month_int
    ).group_by(ShopeeAffCreator.date, ShopeeAffCreator.store_id).all()

    st_conv = db.query(
        func.date(ShopeeAffConversion.order_time),
        ShopeeAffConversion.store_id,
        func.count(ShopeeAffConversion.order_id)
    ).filter(
        func.extract('year',  ShopeeAffConversion.order_time) == year_int,
        func.extract('month', ShopeeAffConversion.order_time) == month_int
    ).group_by(func.date(ShopeeAffConversion.order_time), ShopeeAffConversion.store_id).all()

    matrix = {f"{year_int}-{month_int:02d}-{d:02d}": {} for d in range(1, num_days + 1)}

    def check_presence(records, key_name):
        for r_date, r_store, qty in records:
            if not r_date or not r_store:
                continue
            date_str = r_date.strftime('%Y-%m-%d') if isinstance(r_date, datetime.date) else r_date
            if date_str in matrix:
                if r_store not in matrix[date_str]:
                    matrix[date_str][r_store] = {"product": False, "creator": False, "conversion": False}
                if qty > 0:
                    matrix[date_str][r_store][key_name] = True

    check_presence(st_prod, "product")
    check_presence(st_creator, "creator")
    check_presence(st_conv, "conversion")

    output = [{"date": d, "stores": matrix[d]} for d in sorted(matrix.keys())]
    return {"matrix": output}

@router.delete("/data", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def delete_shopee_data(date: str, store_id: Optional[str] = None, data_type: Optional[str] = None, db: Session = Depends(get_db)):
    """Delete Shopee Affiliate data for a specific date and store. Granular by data_type if provided."""
    del_conv = 0
    del_prod = 0
    del_crtr = 0

    if data_type in (None, 'conversion', 'all'):
        q_conv = db.query(ShopeeAffConversion).filter(func.date(ShopeeAffConversion.order_time) == date)
        if store_id and store_id != 'ALL': q_conv = q_conv.filter(ShopeeAffConversion.store_id == store_id)
        del_conv = q_conv.delete(synchronize_session=False)

    if data_type in (None, 'product', 'all'):
        q_prod = db.query(ShopeeAffProduct).filter(ShopeeAffProduct.date == date)
        if store_id and store_id != 'ALL': q_prod = q_prod.filter(ShopeeAffProduct.store_id == store_id)
        del_prod = q_prod.delete(synchronize_session=False)

    if data_type in (None, 'creator', 'all'):
        q_crtr = db.query(ShopeeAffCreator).filter(ShopeeAffCreator.date == date)
        if store_id and store_id != 'ALL': q_crtr = q_crtr.filter(ShopeeAffCreator.store_id == store_id)
        del_crtr = q_crtr.delete(synchronize_session=False)

    db.commit()

    return {"succeed": True, "message": f"Deleted {del_conv} conversions, {del_prod} products, {del_crtr} creators."}

# ─────────────────────────────────────────────────────────────────────────────
# ANALYTICS (quick top-list)
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/analytics", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def get_analytics(start_date: str, end_date: str, store_id: Optional[str] = None, db: Session = Depends(get_db)):
    base_prod = db.query(
        ShopeeAffProduct.product_id,
        ShopeeAffProduct.product_name,
        func.sum(ShopeeAffProduct.gmv).label('total_gmv')
    ).filter(ShopeeAffProduct.date >= start_date, ShopeeAffProduct.date <= end_date)

    base_creator = db.query(
        ShopeeAffCreator.affiliate_username,
        func.max(ShopeeAffCreator.affiliate_name).label('aff_name'),
        func.sum(ShopeeAffCreator.gmv).label('total_gmv')
    ).filter(ShopeeAffCreator.date >= start_date, ShopeeAffCreator.date <= end_date)

    if store_id and store_id != 'ALL':
        base_prod    = base_prod.filter(ShopeeAffProduct.store_id == store_id)
        base_creator = base_creator.filter(ShopeeAffCreator.store_id == store_id)

    top_products = base_prod.group_by(ShopeeAffProduct.product_id, ShopeeAffProduct.product_name)\
        .order_by(func.sum(ShopeeAffProduct.gmv).desc()).limit(15).all()
    top_creators = base_creator.group_by(ShopeeAffCreator.affiliate_username)\
        .order_by(func.sum(ShopeeAffCreator.gmv).desc()).limit(15).all()

    return {
        "topProducts": [{"product_id": p.product_id, "name": p.product_name, "gmv": p.total_gmv} for p in top_products],
        "topCreators": [{"username": c.affiliate_username, "name": c.aff_name, "gmv": c.total_gmv} for c in top_creators]
    }

# ─────────────────────────────────────────────────────────────────────────────
# FULL REPORT — 3 dimensional JSON
# ─────────────────────────────────────────────────────────────────────────────
def _gmv_aggs():
    return [
        func.sum(case((func.lower(ShopeeAffConversion.order_status) == 'completed', ShopeeAffConversion.purchase_value), else_=0)).label('gmv_completed'),
        func.sum(case((func.lower(ShopeeAffConversion.order_status) == 'pending', ShopeeAffConversion.purchase_value), else_=0)).label('gmv_pending'),
        func.sum(case((func.lower(ShopeeAffConversion.order_status) == 'cancelled', ShopeeAffConversion.purchase_value), else_=0)).label('gmv_canceled'),
    ]

def _build_by_store(db, start_date, end_date, store_id=None):
    """Aggregate by store: GMV (from Conversion) and Units/Clicks (from Creator)."""
    q_conv = db.query(ShopeeAffConversion.store_id, func.sum(ShopeeAffConversion.commission).label('commission'), *_gmv_aggs())\
        .filter(ShopeeAffConversion.order_time >= start_date, ShopeeAffConversion.order_time <= end_date + " 23:59:59")
    if store_id and store_id != 'ALL': q_conv = q_conv.filter(ShopeeAffConversion.store_id == store_id)
    conv_rows = q_conv.group_by(ShopeeAffConversion.store_id).all()

    q_crtr = db.query(ShopeeAffCreator.store_id, func.sum(ShopeeAffCreator.unit_sold).label('units'), func.sum(ShopeeAffCreator.clicks).label('clicks'), func.count(func.distinct(ShopeeAffCreator.affiliate_username)).label('creator_count'))\
        .filter(ShopeeAffCreator.date >= start_date, ShopeeAffCreator.date <= end_date)
    if store_id and store_id != 'ALL': q_crtr = q_crtr.filter(ShopeeAffCreator.store_id == store_id)
    crtr_rows = {r.store_id: r for r in q_crtr.group_by(ShopeeAffCreator.store_id).all()}

    result = []
    for r in conv_rows:
        gmv_c = float(r.gmv_completed or 0)
        gmv_p = float(r.gmv_pending or 0)
        comm = float(r.commission or 0)
        cr = crtr_rows.get(r.store_id)
        result.append({
            "store_id":      r.store_id,
            "gmv_completed": gmv_c, "gmv_pending": gmv_p,
            "gmv_potential": gmv_c + gmv_p,
            "gmv_canceled":  float(r.gmv_canceled or 0),
            "commission":    comm,
            "roi":           round((gmv_c + gmv_p) / comm, 2) if comm > 0 else 0,
            "units":         int(cr.units or 0) if cr else 0,
            "clicks":        int(cr.clicks or 0) if cr else 0,
            "creator_count": int(cr.creator_count or 0) if cr else 0,
        })
    return sorted(result, key=lambda x: x["gmv_potential"], reverse=True)

def _build_by_creator(db, start_date, end_date, store_id=None):
    """Aggregate by creator: GMV (from Conversion) and Units/Clicks (from Creator)."""
    q_conv = db.query(ShopeeAffConversion.affiliate_username, func.max(ShopeeAffConversion.affiliate_name).label('name'), func.sum(ShopeeAffConversion.commission).label('commission'), *_gmv_aggs())\
        .filter(ShopeeAffConversion.order_time >= start_date, ShopeeAffConversion.order_time <= end_date + " 23:59:59")
    if store_id and store_id != 'ALL': q_conv = q_conv.filter(ShopeeAffConversion.store_id == store_id)
    conv_rows = q_conv.group_by(ShopeeAffConversion.affiliate_username).all()

    q_crtr = db.query(ShopeeAffCreator.affiliate_username, func.sum(ShopeeAffCreator.unit_sold).label('units'), func.sum(ShopeeAffCreator.clicks).label('clicks'), func.count(func.distinct(ShopeeAffCreator.store_id)).label('store_count'))\
        .filter(ShopeeAffCreator.date >= start_date, ShopeeAffCreator.date <= end_date)
    if store_id and store_id != 'ALL': q_crtr = q_crtr.filter(ShopeeAffCreator.store_id == store_id)
    crtr_rows = {r.affiliate_username: r for r in q_crtr.group_by(ShopeeAffCreator.affiliate_username).all()}

    result = []
    for r in conv_rows:
        gmv_c = float(r.gmv_completed or 0)
        gmv_p = float(r.gmv_pending or 0)
        comm = float(r.commission or 0)
        cr = crtr_rows.get(r.affiliate_username)
        result.append({
            "username":      r.affiliate_username,
            "name":          r.name or r.affiliate_username,
            "gmv_completed": gmv_c, "gmv_pending": gmv_p,
            "gmv_potential": gmv_c + gmv_p,
            "gmv_canceled":  float(r.gmv_canceled or 0),
            "commission":    comm,
            "roi":           round((gmv_c + gmv_p) / comm, 2) if comm > 0 else 0,
            "units":         int(cr.units or 0) if cr else 0,
            "clicks":        int(cr.clicks or 0) if cr else 0,
            "store_count":   int(cr.store_count or 0) if cr else 0,
        })
    return sorted(result, key=lambda x: x["gmv_potential"], reverse=True)

def _build_by_product(db, start_date, end_date, store_id=None):
    """Aggregate product from Conversion, merge units from Product table, show which creators drove it."""
    q_conv = db.query(
        ShopeeAffConversion.product_id, ShopeeAffConversion.product_name, ShopeeAffConversion.affiliate_username, ShopeeAffConversion.affiliate_name,
        func.sum(ShopeeAffConversion.commission).label('commission'), *_gmv_aggs()
    ).filter(ShopeeAffConversion.order_time >= start_date, ShopeeAffConversion.order_time <= end_date + " 23:59:59")
    if store_id and store_id != 'ALL': q_conv = q_conv.filter(ShopeeAffConversion.store_id == store_id)
    conv_rows = q_conv.group_by(ShopeeAffConversion.product_id, ShopeeAffConversion.product_name, ShopeeAffConversion.affiliate_username, ShopeeAffConversion.affiliate_name).all()

    q_prod = db.query(ShopeeAffProduct.product_id, func.sum(ShopeeAffProduct.unit_sold).label('units'))\
        .filter(ShopeeAffProduct.date >= start_date, ShopeeAffProduct.date <= end_date)
    if store_id and store_id != 'ALL': q_prod = q_prod.filter(ShopeeAffProduct.store_id == store_id)
    prod_units = {r.product_id: int(r.units or 0) for r in q_prod.group_by(ShopeeAffProduct.product_id).all()}

    prod_map = {}
    for r in conv_rows:
        pid = r.product_id
        if pid not in prod_map:
            prod_map[pid] = {
                "product_id": pid, "product_name": r.product_name or "(Unknown)",
                "gmv_completed": 0.0, "gmv_pending": 0.0, "gmv_potential": 0.0, "gmv_canceled": 0.0,
                "commission": 0.0, "units": prod_units.get(pid, 0), "creators": []
            }
        g_c, g_p, g_x = float(r.gmv_completed or 0), float(r.gmv_pending or 0), float(r.gmv_canceled or 0)
        c = float(r.commission or 0)
        prod_map[pid]["gmv_completed"] += g_c
        prod_map[pid]["gmv_pending"]   += g_p
        prod_map[pid]["gmv_potential"] += (g_c + g_p)
        prod_map[pid]["gmv_canceled"]  += g_x
        prod_map[pid]["commission"]    += c
        prod_map[pid]["creators"].append({
            "username": r.affiliate_username,
            "name": r.affiliate_name or r.affiliate_username,
            "gmv_potential": (g_c + g_p), "commission": c,
        })

    result = []
    for pid, p in prod_map.items():
        p["roi"] = round(p["gmv_potential"] / p["commission"], 2) if p["commission"] > 0 else 0
        p["creators"] = sorted(p["creators"], key=lambda x: x["gmv_potential"], reverse=True)[:20]
        p["creator_count"] = len(p["creators"])
        result.append(p)

    result.sort(key=lambda x: x["gmv_potential"], reverse=True)
    return result

@router.get("/report", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def get_full_report(
    start_date: str,
    end_date:   str,
    report_type: str = "by_creator",   # by_store | by_creator | by_product
    store_id:   Optional[str] = None,
    db: Session = Depends(get_db)
):
    if report_type == "by_store":
        return {"data": _build_by_store(db, start_date, end_date, store_id)}
    elif report_type == "by_creator":
        return {"data": _build_by_creator(db, start_date, end_date, store_id)}
    elif report_type == "by_product":
        return {"data": _build_by_product(db, start_date, end_date, store_id)}
    else:
        raise HTTPException(status_code=400, detail="Invalid report_type")

# ─────────────────────────────────────────────────────────────────────────────
# DOWNLOAD EXCEL — professional styled workbook
# ─────────────────────────────────────────────────────────────────────────────
def _build_excel(data, report_type, period_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header style constants
    HDR_FILL  = PatternFill("solid", fgColor="1A3A5C")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
    THIN      = Side(style="thin", color="2D5A8E")
    THIN_GRAY = Side(style="thin", color="CBD5E1")
    HDR_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CELL_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    NUM_ALIGN = Alignment(horizontal="right")
    ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")

    def write_header(ws, headers, row=1):
        ws.row_dimensions[row].height = 32
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill   = HDR_FILL
            cell.font   = HDR_FONT
            cell.border = HDR_BORDER
            cell.alignment = HDR_ALIGN

    def apply_cell(ws, row, col, value, is_alt=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = CELL_BORDER
        if is_alt:
            cell.fill = ALT_FILL
        if isinstance(value, (int, float)):
            cell.alignment = NUM_ALIGN
        return cell

    def fmt_rp(v):
        return v  # Store as number, format in Excel number format

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Shopee Affiliate Report — {report_type.replace('_', ' ').upper()} | {period_label}"
    title_cell.font  = Font(bold=True, size=13, color="1A3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── By Store ──────────────────────────────────────────────────────────────
    if report_type == "by_store":
        headers = ["No", "Store ID", "GMV Completed", "GMV Pending", "GMV Potential", "GMV Canceled", "Commission (Rp)", "ROI", "Units Sold", "Clicks", "Total Creators"]
        write_header(ws, headers, row=3)
        for i, r in enumerate(data, 1):
            alt = (i % 2 == 0)
            apply_cell(ws, i+3, 1, i,                alt)
            apply_cell(ws, i+3, 2, r["store_id"],    alt)
            apply_cell(ws, i+3, 3, r["gmv_completed"],alt).number_format = '#,##0'
            apply_cell(ws, i+3, 4, r["gmv_pending"],  alt).number_format = '#,##0'
            apply_cell(ws, i+3, 5, r["gmv_potential"],alt).number_format = '#,##0'
            apply_cell(ws, i+3, 6, r["gmv_canceled"], alt).number_format = '#,##0'
            apply_cell(ws, i+3, 7, r["commission"],   alt).number_format = '#,##0'
            apply_cell(ws, i+3, 8, r["roi"],          alt).number_format = '0.00'
            apply_cell(ws, i+3, 9, r["units"],        alt)
            apply_cell(ws, i+3, 10, r["clicks"],      alt)
            apply_cell(ws, i+3, 11, r["creator_count"], alt)
        col_widths = [5, 18, 18, 18, 18, 18, 18, 10, 15, 12, 16]

    # ── By Creator ────────────────────────────────────────────────────────────
    elif report_type == "by_creator":
        headers = ["No", "Username", "Creator Name", "GMV Completed", "GMV Pending", "GMV Potential", "GMV Canceled", "Commission (Rp)", "ROI", "Units Sold", "Clicks", "Total Stores"]
        write_header(ws, headers, row=3)
        for i, r in enumerate(data, 1):
            alt = (i % 2 == 0)
            apply_cell(ws, i+3, 1, i,               alt)
            apply_cell(ws, i+3, 2, r["username"],   alt)
            apply_cell(ws, i+3, 3, r["name"],       alt)
            apply_cell(ws, i+3, 4, r["gmv_completed"],alt).number_format = '#,##0'
            apply_cell(ws, i+3, 5, r["gmv_pending"],  alt).number_format = '#,##0'
            apply_cell(ws, i+3, 6, r["gmv_potential"],alt).number_format = '#,##0'
            apply_cell(ws, i+3, 7, r["gmv_canceled"], alt).number_format = '#,##0'
            apply_cell(ws, i+3, 8, r["commission"], alt).number_format = '#,##0'
            apply_cell(ws, i+3, 9, r["roi"],        alt).number_format = '0.00'
            apply_cell(ws, i+3, 10, r["units"],      alt)
            apply_cell(ws, i+3, 11, r["clicks"],     alt)
            apply_cell(ws, i+3, 12, r["store_count"],alt)
        col_widths = [5, 24, 28, 18, 18, 18, 18, 18, 10, 14, 12, 13]

    # ── By Product ────────────────────────────────────────────────────────────
    elif report_type == "by_product":
        # Sheet 1: product summary
        headers = ["No", "Product ID", "Product Name", "GMV Completed", "GMV Pending", "GMV Potential", "GMV Canceled", "Commission (Rp)", "ROI", "Units Sold", "Total Creators"]
        write_header(ws, headers, row=3)
        for i, r in enumerate(data, 1):
            alt = (i % 2 == 0)
            apply_cell(ws, i+3, 1, i,                  alt)
            apply_cell(ws, i+3, 2, r["product_id"],    alt)
            apply_cell(ws, i+3, 3, r["product_name"],  alt)
            apply_cell(ws, i+3, 4, r["gmv_completed"], alt).number_format = '#,##0'
            apply_cell(ws, i+3, 5, r["gmv_pending"],   alt).number_format = '#,##0'
            apply_cell(ws, i+3, 6, r["gmv_potential"], alt).number_format = '#,##0'
            apply_cell(ws, i+3, 7, r["gmv_canceled"],  alt).number_format = '#,##0'
            apply_cell(ws, i+3, 8, r["commission"],    alt).number_format = '#,##0'
            apply_cell(ws, i+3, 9, r["roi"],           alt).number_format = '0.00'
            apply_cell(ws, i+3, 10, r["units"],        alt)
            apply_cell(ws, i+3, 11, r["creator_count"], alt)
        col_widths = [5, 18, 50, 18, 18, 18, 18, 18, 10, 14, 16]

        # Sheet 2: product × creator breakdown
        ws2 = wb.create_sheet("Creator per Produk")
        ws2.merge_cells("A1:G1")
        ws2["A1"].value = "Detail Creator per Produk"
        ws2["A1"].font  = Font(bold=True, size=13, color="1A3A5C")
        ws2["A1"].alignment = Alignment(horizontal="center")
        ws2.row_dimensions[1].height = 28
        det_headers = ["Product Name", "Username Creator", "Creator Name", "GMV Potential (Rp)", "Commission (Rp)", "ROI"]
        write_header(ws2, det_headers, row=3)
        row_idx = 4
        alt = False
        for p in data:
            for c in p["creators"]:
                comm = c["commission"]
                roi  = round(c["gmv_potential"] / comm, 2) if comm > 0 else 0
                apply_cell(ws2, row_idx, 1, p["product_name"], alt)
                apply_cell(ws2, row_idx, 2, c["username"],     alt)
                apply_cell(ws2, row_idx, 3, c["name"],         alt)
                apply_cell(ws2, row_idx, 4, c["gmv_potential"],alt).number_format = '#,##0'
                apply_cell(ws2, row_idx, 5, c["commission"],   alt).number_format = '#,##0'
                apply_cell(ws2, row_idx, 6, roi,               alt).number_format = '0.00'
                row_idx += 1
                alt = not alt
        for ci, w in enumerate([50, 24, 28, 16, 16, 10], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

    # Apply column widths to main sheet
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze panes
    ws.freeze_panes = ws["A4"]

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/report/download", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def download_report(
    start_date:  str,
    end_date:    str,
    report_type: str = "by_creator",
    store_id:    Optional[str] = None,
    db: Session = Depends(get_db)
):
    if report_type == "by_store":
        data = _build_by_store(db, start_date, end_date, store_id)
    elif report_type == "by_creator":
        data = _build_by_creator(db, start_date, end_date, store_id)
    elif report_type == "by_product":
        data = _build_by_product(db, start_date, end_date, store_id)
    else:
        raise HTTPException(status_code=400, detail="Invalid report_type")

    period_label = f"{start_date} s.d. {end_date}"
    excel_buf    = _build_excel(data, report_type, period_label)
    filename     = f"Shopee_Affiliate_{report_type}_{start_date}_{end_date}.xlsx"

    return StreamingResponse(
        excel_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ─────────────────────────────────────────────────────────────────────────────
# COMPARISON — Period A vs Period B
# ─────────────────────────────────────────────────────────────────────────────
def _aggregate_for_comparison(db, start_date, end_date, dimension, store_id=None):
    """Returns dict keyed by entity-id for easy merging."""
    if dimension == "by_store":
        rows = _build_by_store(db, start_date, end_date, store_id)
        return {r["store_id"]: r for r in rows}
    elif dimension == "by_creator":
        rows = _build_by_creator(db, start_date, end_date, store_id)
        return {r["username"]: r for r in rows}
    elif dimension == "by_product":
        rows = _build_by_product(db, start_date, end_date, store_id)
        return {r["product_id"]: r for r in rows}
    return {}

@router.get("/comparison", dependencies=[Depends(require_tool_access("affiliate_performance"))])
def get_comparison(
    period_a_start: str,
    period_a_end:   str,
    period_b_start: str,
    period_b_end:   str,
    dimension:      str = "by_creator",   # by_store | by_creator | by_product
    store_id:       Optional[str] = None,
    db: Session = Depends(get_db)
):
    a_data = _aggregate_for_comparison(db, period_a_start, period_a_end, dimension, store_id)
    b_data = _aggregate_for_comparison(db, period_b_start, period_b_end, dimension, store_id)

    all_keys = sorted(set(list(a_data.keys()) + list(b_data.keys())))

    def delta(a_val, b_val):
        a = a_val or 0
        b = b_val or 0
        if b == 0:
            return None
        return round(((a - b) / b) * 100, 1)

    def label_for(key, a_data, b_data, dimension):
        r = a_data.get(key) or b_data.get(key)
        if dimension == "by_store":   return r["store_id"]
        if dimension == "by_creator": return f"{r.get('name', '')} (@{r['username']})"
        if dimension == "by_product": return f"{r.get('product_name', key)} (PID: {key})"
        return key

    rows = []
    for key in all_keys:
        a = a_data.get(key, {})
        b = b_data.get(key, {})

        a_gmv = a.get("gmv_potential", 0)
        b_gmv = b.get("gmv_potential", 0)
        a_comm = a.get("commission", 0)
        b_comm = b.get("commission", 0)
        a_roi = round(a_gmv / a_comm, 2) if a_comm > 0 else 0
        b_roi = round(b_gmv / b_comm, 2) if b_comm > 0 else 0

        row = {
            "key":          key,
            "label":        label_for(key, a_data, b_data, dimension),
            "a_gmv":        a_gmv,
            "b_gmv":        b_gmv,
            "delta_gmv":    delta(a_gmv, b_gmv),
            "a_commission": a_comm,
            "b_commission": b_comm,
            "delta_commission": delta(a_comm, b_comm),
            "a_roi":        a_roi,
            "b_roi":        b_roi,
            "delta_roi":    delta(a_roi, b_roi),
            "a_units":      a.get("units", 0),
            "b_units":      b.get("units", 0),
            "delta_units":  delta(a.get("units", 0), b.get("units", 0)),
            "a_clicks":     a.get("clicks", 0),
            "b_clicks":     b.get("clicks", 0),
            "delta_clicks": delta(a.get("clicks", 0), b.get("clicks", 0)),
        }
        rows.append(row)

    # Sort by period A GMV descending (most important first)
    rows.sort(key=lambda x: x["a_gmv"], reverse=True)

    return {
        "rows":     rows,
        "period_a": f"{period_a_start} → {period_a_end}",
        "period_b": f"{period_b_start} → {period_b_end}",
    }
